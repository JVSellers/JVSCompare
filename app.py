
import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from copy import copy
import re

st.set_page_config(page_title="Amazon Product Loader", layout="wide")
st.image("logo.jpeg", width=120)
st.title("Amazon Product Loader - JVSellers")

st.markdown("### 1. Subir archivo Excel existente")
uploaded_file = st.file_uploader("Selecciona tu archivo Excel (.xlsm)", type=["xlsm"])

if "temp_table" not in st.session_state:
    st.session_state.temp_table = pd.DataFrame()

if uploaded_file:
    bytes_data = uploaded_file.read()
    wb = load_workbook(filename=BytesIO(bytes_data), data_only=False, keep_vba=True)
    sheet_alta = wb["Alta de productos"]
    sheet_calc = wb["calc. precio minimo intern"]

    st.markdown("### Productos existentes:")
    data = list(sheet_alta.values)
    headers = data[0]
    df_alta = pd.DataFrame(data[1:], columns=headers)
    st.dataframe(df_alta, use_container_width=True)

    st.markdown("---")
    st.markdown("### 2. A√±adir nuevo producto desde URL de Amazon")
    url = st.text_input("Pega aqu√≠ la URL de Amazon")

    def obtener_info_amazon(url):
        headers = {"User-Agent": "Mozilla/5.0"}
        try:
            res = requests.get(url, headers=headers, timeout=10)
            soup = BeautifulSoup(res.text, "html.parser")

            title_tag = soup.select_one("#productTitle")
            title = title_tag.get_text(strip=True) if title_tag else "Nombre no encontrado"

            price_tag = soup.select_one("span.a-price > span.a-offscreen")
            price = price_tag.get_text(strip=True) if price_tag else "Precio no disponible"

            asin = None
            if "/dp/" in url:
                asin = url.split("/dp/")[1].split("?")[0]
            elif "asin=" in url:
                asin = url.split("asin=")[1].split("&")[0]

            prime = bool(soup.select_one("i[aria-label*='Prime']"))
            rating_tag = soup.select_one("span.a-icon-alt")
            rating = rating_tag.get_text(strip=True) if rating_tag else "N/A"

            return {
                "Nombre del Articulo": title,
                "ASIN": asin,
                "Precio": price,
                "PRIMEABLE": "S√≠" if prime else "No",
                "Valoraci√≥n": rating,
                "Url del producto": url
            }
        except Exception as e:
            return {"Error": str(e)}

    if st.button("A√±adir producto"):
        if url:
            nuevo_producto = obtener_info_amazon(url)
            if "Error" not in nuevo_producto:
                st.session_state.temp_table = pd.concat(
                    [st.session_state.temp_table, pd.DataFrame([nuevo_producto])],
                    ignore_index=True
                )
                st.success("‚úÖ Producto a√±adido a la tabla temporal")
            else:
                st.error("Error al obtener informaci√≥n: " + nuevo_producto["Error"])
        else:
            st.warning("Por favor, introduce una URL v√°lida.")

    st.markdown("### Productos a√±adidos en esta sesi√≥n:")
    st.dataframe(st.session_state.temp_table, use_container_width=True)

    if st.button("Eliminar √∫ltimo producto"):
        if not st.session_state.temp_table.empty:
            st.session_state.temp_table = st.session_state.temp_table.iloc[:-1]

    st.markdown("---")
    if st.button("Descargar Excel actualizado"):

        def adjust_formula(formula, old_row, new_row):
            def repl(m):
                col = m.group(1)
                row = int(m.group(2))
                return f"{col}{row + (new_row - old_row)}"
            return re.sub(r"([A-Z]{1,3})(\d+)", repl, formula)

        def clone_row(ws, row_idx):
            new_idx = row_idx + 1
            ws.insert_rows(new_idx)
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                new_cell = ws.cell(row=new_idx, column=col_idx)
                if cell.data_type == 'f' and isinstance(cell.value, str):
                    new_cell.value = adjust_formula(cell.value, row_idx, new_idx)
                else:
                    new_cell.value = cell.value
                if cell.has_style:
                    new_cell._style = copy(cell._style)
                if cell.hyperlink:
                    new_cell.hyperlink = copy(cell.hyperlink)

        # Ampliar tabla si existe
        if sheet_alta._tables:
            table = list(sheet_alta._tables.values())[0]
            ref = table.ref
            start_col_letter = ref.split(":")[0][0]
            start_row = int(ref.split(":")[0][1:])
            end_col_letter = ref.split(":")[1][0]
            end_row = int(ref.split(":")[1][1:])
            new_end_row = end_row + len(st.session_state.temp_table)
            table.ref = f"{start_col_letter}{start_row}:{end_col_letter}{new_end_row}"

        last_row_alta = sheet_alta.max_row
        for _, row in st.session_state.temp_table.iterrows():
            clone_row(sheet_alta, last_row_alta)
            last_row_alta += 1
            sheet_alta.cell(row=last_row_alta, column=2).value = row["Nombre del Articulo"]
            sheet_alta.cell(row=last_row_alta, column=2).hyperlink = row["Url del producto"]
            sheet_alta.cell(row=last_row_alta, column=2).style = "Hyperlink"
            st.success(f"Producto a√±adido en fila {last_row_alta} de 'Alta de productos'")

        last_row_calc = sheet_calc.max_row
        for _, row in st.session_state.temp_table.iterrows():
            clone_row(sheet_calc, last_row_calc)
            last_row_calc += 1
            sheet_calc.cell(row=last_row_calc, column=1).value = row["Nombre del Articulo"]
            sheet_calc.cell(row=last_row_calc, column=1).hyperlink = row["Url del producto"]
            sheet_calc.cell(row=last_row_calc, column=1).style = "Hyperlink"
            sheet_calc.cell(row=last_row_calc, column=2).value = "SI"

        output = BytesIO()
        wb.save(output)
        st.download_button("üì• Descargar archivo Excel", data=output.getvalue(), file_name="productos_actualizados.xlsm")
